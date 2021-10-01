import openpyxl
import os
import re
import pyperclip
import pandas as pd
from time import sleep
import datetime
from alive_progress import alive_bar

#Change directory for your xlsx file
os.chdir('c:\\Users\\alexp\\Downloads')

now = datetime.datetime.now()
text = 'Welcome to my learning program!\n\tThis program is made for a better understanding on how openpyxl works.'
if now.hour < 12:
    print('Good morning !\n\t',text)
elif now.hour > 12:
    print('Good afternoon !\n\t',text)
else:
    print('Good evening !\n\t',text)

#Open the xlsx file
excel_name = input('Enter excel name: ')
try:
    workbook = openpyxl.load_workbook(excel_name + '.xlsx')
except:
    print("The file doesn't exist or it's not in the selected path.")
    quit()
show_xlsx_data = workbook.get_sheet_names()
# Show the sheets of xlsx file
print('Xlsx info: ', show_xlsx_data)

choose = input('What sheet do you want to work with?\nSheet name: ')
#Show what the sheet is containing
try:
    sheet = workbook.get_sheet_by_name(choose)
    for row in sheet.rows:
        for cell in row:
            print(row, cell.value)
except Exception as error:
    print(error)

command_list = ['Show column', 'Show rows','Show cell value','Append','Indetify a cell','Make me a sheet','Change xlsx file']
print('List of commands:', command_list)

def XLSX_sandbox():
    command = input('Enter a command: ')
    while True:
        if command == 'Change xlsx file':

            def cd(): #Enter a directory and call the xlsx file. As you call the xlsx file you will get the data
                      # inside the xlsx file and after this you can call different functions
                cd = input('Enter the xlsx file path: ')
                os.chdir(cd)
                print('The directory has been changed:', os.path)
                excel_name = input('Enter the xlsx file name you are looking for in this directory: ')
                workbook = openpyxl.load_workbook(excel_name + '.xlsx')
                show_xlsx_data = workbook.get_sheet_names()
                print('Xlsx info: ', show_xlsx_data)
                choose = input('What sheet do you want to work with?\nSheet name: ')
                sheet = workbook.get_sheet_by_name(choose)
                for row in sheet.rows:
                    for cell in row:
                        print(row, cell.value)
                XLSX_sandbox()
            cd()

        if command == 'Show column':

            def show_column(): #For a specified xlsx file you can see the data from a desired column.
                               # The input for this function is the name of the column. For example: A, B, AA etc
                n_column = input('What column would you like to see?\nColumn: ')
                if n_column == 'Back':
                    XLSX_sandbox()
                else:
                    column = sheet[n_column]
                    for cell in column:
                        print(cell ,cell.value)
            show_column()

        elif command == 'Show row':

            def show_row():#For a specified xlsx file you can see the data from a desired row.
                           # The input for this function is the number of the row. For example: 2, 3, 124 etc
                n_row = input('What row would you like to see?\nRow: ')
                if n_row == 'Back':
                    XLSX_sandbox()
                else:
                    row = sheet[n_row]
                    for cell in row:
                        print(cell.value)
            show_row()

        elif command == 'Show a cell':

            def show_a_cell():#For a specified xlsx file you can see the data from a desired cell.
                              # The input for this function is the number of the row and column.
                              # For example: Column X = 2
                              #              Row Y = 2
                print('If you want to go back to command input just press enter for X and Y.')
                X = input('Column position: ')
                Y = input('Row position: ')
                if X.isdigit() and Y.isdigit():
                    cell = sheet.cell(row=int(Y), column=int(X))
                    print('Cell value:', cell.value)
                elif X or Y == '':
                    XLSX_sandbox()
            show_a_cell()

        elif command == 'Make me a sheet':

            def make_a_sheet():#For a specified xlsx file you can create a sheet. After creating a new sheet you can
                               #either append data with pyperclip or read data.
                               #As you create a new sheet this program will create another xlsx file in the same directory with the name of "example2".xlsx
                name = input("What's the title you want to give to the sheet?\nIf you just want to read a specific sheet enter '1'\nTitle: ")
                if len(name) < 1:
                    XLSX_sandbox()
                elif name == str(1):
                    print()
                    sheet_name = input('Enter the sheet name you want to read: ')
                    sheet = workbook.get_sheet_by_name(sheet_name)
                    for row in sheet.rows:
                        for cell in row:
                            print(row, cell.value)
                else:
                    sheet_add = workbook.create_sheet(index=0,title=name)
                    print(sheet_add)
                    print('The sheet have been created!')
                    print("The title you've added is:", sheet_add.title)
                    print('The new list of sheets:', workbook.get_sheet_names())
                    workbook.save('example2.xlsx')
                    sheet_command_list = ['Append data ( Using pyperclip-paste clipboard )', 'Read data']
                    print('Do you want me to execute some of these commands to this sheet?\n', sheet_command_list)
                    ex = input('Enter command or enter "No".\n')
                    if ex == 'No':
                        XLSX_sandbox()
                    elif ex == 'Append data':
                        # creating a regex for phone numbers
                        phone_regex = re.compile(r'''
                        # 315-123-2143, 444-0000, (321) 455-0994
                        (
                        (\d\d\d | (\d\d\d) | (\(\d\d\d))?        # are code ( optional )
                        (\s|-)                  # first separator
                        \d\d\d                  # first 3 digits
                        -                       # separator
                        \d\d\d\d                # last 4 digits
                        )
                        ''', re.VERBOSE)
                        #get the date from clipboard and make a list with the collected data
                        data = pyperclip.paste()
                        extract_phone_numbers = phone_regex.findall(data)
                        phone_num_list = list()
                        for ph_number in extract_phone_numbers:
                            phone_num_list.append(ph_number[0])
                        print(phone_num_list)
                        # for every phone number in phone_num_list append data in the excel sheet you've just named
                        with alive_bar(len(phone_num_list), bar='bubbles', spinner='notes2') as bar:
                            for i in range(len(phone_num_list)):
                                data_frame = pd.DataFrame({'Phone numbers': phone_num_list})
                                writer = pd.ExcelWriter('example2.xlsx')
                                data_frame.to_excel(writer, sheet_name=name, index=False)
                                sleep(0.03)
                                bar()
                                writer.save()
                        print("The data has been added!\nATENTION!\n\tThe new sheet is added in a copy of the initial xlsx file with the same path.")

                    elif ex == 'Read data':
                        sheet_name = input('Enter the name of the sheet: ')
                        sheet = workbook.get_sheet_by_name(sheet_name)
                        for row in sheet.rows:
                            for cell in row:
                                print(row, cell.value)
                    question = input("Would you like to do something else? [YES/NO]\n")
                    if question == 'NO':
                        XLSX_sandbox()
                    elif question == 'YES':
                        make_a_sheet()
            make_a_sheet()

        elif command == 'Edit a cell':

            def indetify(sheet):#This function will print the desired cell coordinate and it's value by simply adding the direct
                                #coordinate. Input example: A2, A5, R33
                coord = input('What is the coordinate for the cell you are looking for?\nInput examples: D5, A1, C2 (Column+Row)\nEnter cell coordonate: ')
                if coord == 'Back':
                    XLSX_sandbox()
                else:
                    cell_indetify = sheet[coord]
                    print('\nThe cell coordonate is:', cell_indetify,'\nThe cell value is:',cell_indetify.value,'\n')
            indetify(sheet)

        elif command == 'Stop':
            workbook.close()
            print('Thank you and have a nice day !')

XLSX_sandbox()
